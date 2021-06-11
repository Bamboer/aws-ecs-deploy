#! /usr/local/env python3
#codind:utf-8
import os
import json
from urllib import request
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import SheetFormatProperties
from openpyxl.styles import Font,Alignment,Side,Border,PartternFill

GpuNode = "gpu6803|gpu6807|10.11.3.1|gpu6806|gpu6870|192.168.68.85|192.168.3.4|192.168.68.2|192.168.68.4|192.168.68.7|10.12.3.1"

PrdMemRelurl = "http://prom.devops.prd.dm-ai.cn/api/v1/query?" +urlencode({"query":'sum(container_memory_working_set_bytes{pod!="",image!="",node!~"%s",namespace!~"apollo|kube-system|kubernetes-dashboard"}) by(namespace)'%GpuNode})
PrdMemRequrl = "http://prom.devops.prd.dm-ai.cn/api/v1/query?" +urlencode({"query":'sum(kube_pod_container_resource_requests_memory_bytes{pod!="",node!~"%s",namespace!~"apollo|kube-system|kubernetes-dashboard"}) by (namespace)'%GpuNode})
PrdCpuRelurl = "http://prom.devops.prd.dm-ai.cn/api/v1/query?" +urlencode({"query":'sum(rate(container_cpu_usage_seconds_total{pod!="",image!="",node!~"%s",namespace!~"apollo|kube-system|kubernetes-dashboard"}[5m])) by(namespace)'%GpuNode})
PrdCpuRequrl = "http://prom.devops.prd.dm-ai.cn/api/v1/query?" +urlencode({"query":'sum(kube_pod_container_resource_requests_cpu_cores{pod!="",node!~"%s",namespace!~"apollo|kube-system|kubernetes-dashboard"}) by(namespace)'%GpuNode})
PrdGpuRelurl = "http://prom.devops.prd.dm-ai.cn/api/v1/query?" +urlencode({"query":'sum(pod_used_gpu_mem_MB * 1024 * 1024) by (namespace)'})
PrdGpuRequrl = "http://prom.devops.prd.dm-ai.cn/api/v1/query?" +urlencode({"query":'sum(kube_pod_container_resource_limits{resource="aliyun_com_gpu_mem",namespace!~"sci-demo|devops|smtc"} * 1024 * 1024 ) by(namespace)'})


DevMemRelurl = ""
DevMemReqrul = ""
DevCpuRelurl = ""
DevCpuRequrl = ""
DevGpuRelurl = ""
DevGpuRequrl = ""

StageMemRelurl = "" 
StageMemReqrul = ""
StageCpuRelurl = ""
StageCpuRequrl = ""
StageGpuRelurl = ""
StageGpuRequrl = ""

TestMemRelurl = ""
TestMemReqrul = ""
TestCpuRelurl = ""
TestCpuRequrl = ""
TestGpuRelurl = ""
TestGpuRequrl = ""

UatMemRelurl = ""
UatMemReqrul = ""
UatCpuRelurl = ""
UatCpuRequrl = ""
UatGpuRelurl = ""
UatGpuRequrl = ""

PrdImgUrl = "http://grafana.ops.dm-ai.cn/d/k8s-pod-zhanshi/k8s-podzi-yuan-li-yong-zhan-shi?orgId=1&var-Datasource=k8s-prd&var-Namespace=%s&var-Deployment=All&var-Statefulset=All&var-Daemonset=All&var-Pod=All&var-interval=$__auto_interval_interval&var-Node=All"%x
DevImgUrl = "http://grafana.ops.dm-ai.cn/d/k8s-pod-zhanshi/k8s-podzi-yuan-li-yong-zhan-shi?orgId=1&var-Datasource=k8s-prd&var-Namespace=%s&var-Deployment=All&var-Statefulset=All&var-Daemonset=All&var-Pod=All&var-interval=$__auto_interval_interval&var-Node=All"%x
StageImgUrl = "http://grafana.ops.dm-ai.cn/d/k8s-pod-zhanshi/k8s-podzi-yuan-li-yong-zhan-shi?orgId=1&var-Datasource=k8s-prd&var-Namespace=%s&var-Deployment=All&var-Statefulset=All&var-Daemonset=All&var-Pod=All&var-interval=$__auto_interval_interval&var-Node=All"%x
TestImgUrl = "http://grafana.ops.dm-ai.cn/d/k8s-pod-zhanshi/k8s-podzi-yuan-li-yong-zhan-shi?orgId=1&var-Datasource=k8s-prd&var-Namespace=%s&var-Deployment=All&var-Statefulset=All&var-Daemonset=All&var-Pod=All&var-interval=$__auto_interval_interval&var-Node=All"%x
UatImgUrl = "http://grafana.ops.dm-ai.cn/d/k8s-pod-zhanshi/k8s-podzi-yuan-li-yong-zhan-shi?orgId=1&var-Datasource=k8s-prd&var-Namespace=%s&var-Deployment=All&var-Statefulset=All&var-Daemonset=All&var-Pod=All&var-interval=$__auto_interval_interval&var-Node=All"%x


URL = {"PRD":{"ImgUrl": PrdImgUrl,"mem":[PrdMemRelurl,PrdMemRequrl],"cpu":[PrdCpuRelurl,PrdCpuRequrl],"gpu":[PrdGpuRelurl,PrdGpuRequrl]},"DEV":{"ImgUrl": DevImgUrl,"mem":[DevMemRelurl,DevMemRequrl],"cpu":[DevCpuRelurl,DevCpuRequrl],"gpu":[DevGpuRelurl,DevGpuRequrl]},"STAGE":{"ImgUrl":StageImgUrl,"mem":[StageMemRelurl,StageMemRequrl],"cpu":[StageCpuRelurl,StageCpuRequrl],"gpu":[StageGpuRelurl,StageGpuRequrl]},"TEST":{"ImgUrl": TestImgUrl,"mem":[TestMemRelurl,TestMemRequrl],"cpu":[TestCpuRelurl,TestCpuRequrl],"gpu":[TestGpuRelurl,TestGpuRequrl]},"UAT":{"ImgUrl": UatImgUrl,"mem":[UatMemRelurl,UatMemRequrl],"cpu":[UatCpuRelurl,UatCpuRequrl],"gpu":[UatGpuRelurl,UatGpuRequrl]}}
prd = {"mem":{},"cpu":{},"gpu":{}}
dev = {"mem":{},"cpu":{},"gpu":{}}
stage = {"mem":{},"cpu":{},"gpu":{}}
test ={"mem":{},"cpu":{},"gpu":{}}
uat = {"mem":{},"cpu":{},"gpu":{}}
D = {"PRD": prd,"DEV": dev,"STAGE": stage,"TEST": test,"UAT": uat}

font = Font(name='微软雅黑',size=20,bold=True,italic=Treu,color='White')
alignment = Alignment(horizontal='center',vertical='center',text_roation=45,wrap_text=True)
side = Side(style='thin',color='FF000000')
border = Border(left=side,right=side,top=side,bottom=side)
pattern_fill1 = PartternFill(fill_type='solid',fgColor='black')
pattern_fill2 = PartternFill(fill_type='solid',fgColor='99ccff')
wb = Workbook()
excel_name = os.path.join(os.getcwd(),"daily_export.xlsx")

def GetData(url):
	M = {}
    req = request.Request(url,method='GET')
    try:
        r = request.urlopen(req)
        json_data = json.loads(r.read().decode('utf-8')) 
        result = json_data['data']['result']
        if json_data['status'] != 'success':
        	raise Exception('request data failure.')
        for i in result:
        	M[i['metric']['namespace']] = i['value'][1]
        return M 
    except Exception as e:
        raise Exception(e)
def Data(d1,d2,sourcename,enviroment):
	P ={}
    for i in d1:
        if i in d2.keys():
            if i not in P.keys():
                P[i] = []
            if sourcename != 'cpu':
                P[i] = [i,sourcename,round(int(d2[i])/1024/1024/1024,2),round(int(d1[i])/1024/1024/1024,2),str(round(float(d1[i])/float(d2[i])*100,2))+'%','']
            else:
                P[i] = [i,sourcename,round(float(d2[i]),3),round(float(d1[i]),3),str(round(float(d1[i])/float(d2[i])*100,2))+'%','']
        else:
            print(enviroment,">",sourcename,":",i ,"relize not in request.")
    return P         
def GeneratorExcel():
	for enviroment in URL.keys():
		for resourcename in D[enviroment].keys():
			d1 = {}
			d2 = {}
			for length in range(len(URL[enviroment][resourcename])):
				if length == 0:
					d1 = GetData(URL[enviroment][resourcename][length])
				if length == 1:
					d2 = GetData(URL[enviroment][resourcename][length])
			D[enviroment][resourcename] = Data(d1,d2,sourcename,enviroment)
	for enviroment in D.keys():
		ws = wb.create_sheet()
		ws.title = enviroment
        ws.sheet_format = SheetFormatProperties(defaultColWidth=18.0,defaultRowHeight=30.0)
        tableTitle = ['产品名称', '资源类别', '请求值', '使用值','使用率','备注','','产品名称','资源类别','请求值(GB)','使用值(GB)','使用率','备注','','产品名称','资源类别','请求值(GB)','使用值(GB)','使用率','备注','']
        for col in range(len(tableTitle)):
        	c = col+1
        	ws.cell(row=1,column=c).value = tableTitle[col]
        	if 21%c == 0:
        		continue
        	ws.cell(row=1,column=c).border = border
        	ws.cell(row=1,column=c).alignment = alignment
        	ws.cell(row=1,column=c).fill = pattern_fill1
        	ws.cell(row=1,column=c).font = font 
        for r in range(len(D[enviroment]["cpu"])):
        	for service in D[enviroment]["cpu"]:
        		for i in range(len(D[enviroment]["cpu"][service])):
        		    ws.cell(row=r+2,column=i+1).value = D[enviroment]["cpu"][service][i]
        		    ws.cell(row=r+2,column=i+1).border = border
        		    ws.cell(row=r+2,column=i+1).alignment = alignment
        		    if i == 0:
        		    	ws.cell(row=r+2,column=i+1).hyperlink = URL[enviroment]["ImgUrl"]%(service)
        		    if i in [0,1]:
        		        ws.cell(row=r+2,column=i+1).fill = pattern_fill2
        for r in range(len(D[enviroment]["mem"])):
        	for service in D[enviroment]["mem"]:
        		for i in range(len(D[enviroment]["mem"][service])):
        		    ws.cell(row=r+2,column=i+8).value = D[enviroment]["mem"][service][i]
        		    ws.cell(row=r+2,column=i+8).border = border
        		    ws.cell(row=r+2,column=i+8).alignment = alignment
        		    if i == 0:
        		    	ws.cell(row=r+2,column=i+7).hyperlink = URL[enviroment]["ImgUrl"]%(service)
        		    if i in [0,1]:
        		        ws.cell(row=r+2,column=i+7).fill = pattern_fill2
        for r in range(len(D[enviroment]["gpu"])):
        	for service in D[enviroment]["gpu"]:
        		for i in range(len(D[enviroment]["gpu"][service])):
        		    ws.cell(row=r+2,column=i+15).value = D[enviroment]["gpu"][service][i]
        		    ws.cell(row=r+2,column=i+15).border = border
        		    ws.cell(row=r+2,column=i+15).alignment = alignment
        		    if i == 0:
        		    	ws.cell(row=r+2,column=i+14).hyperlink = URL[enviroment]["ImgUrl"]%(service)
        		    if i in [0,1]:
        		        ws.cell(row=r+2,column=i+14).fill = pattern_fill2
    wb.save(filename=excel_name)

GeneratorExcel()    